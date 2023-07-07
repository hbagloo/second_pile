import numpy as np
from openpyxl import *
import matplotlib.pyplot as plt

pile_name = "Pile #918"
wb = load_workbook("pile918.xlsx")
ws = wb.worksheets[0]

# 1 - definitions
sample_number_tuple = ()
section_tuple = ()
subrow_tuple = ()
location_tuple = ()

source_list = []
number_of_sections = 27

# 2 - reading data from Excel
for row in range(1, ws.max_row + 1):
    for column in range(1, ws.max_column + 1):
        if ws.cell(row=row, column=column).value == "Locations":
            location_row = row
            location_column = column
        if ws.cell(row=row, column=column).value == "سکشن":
            section_row = row
            section_column = column
        if ws.cell(row=row, column=column).value == "رج":
            subrow_row = row
            subrow_column = column
        if ws.cell(row=row, column=column).value == "Pile Code" or \
                ws.cell(row=row, column=column).value == "شماره نمونه":
            sample_number_row = row
            sample_number_column = column

for row in range(section_row + 1, ws.max_row + 1):
    if ws.cell(row=row, column=sample_number_column).value != "None":
        sample_number_tuple += (ws.cell(row=row, column=sample_number_column).value,)
    if ws.cell(row=row, column=section_column).value != "None":
        section_tuple += (ws.cell(row=row, column=section_column).value,)
    if ws.cell(row=row, column=section_column).value != "None":
        subrow_tuple += (ws.cell(row=row, column=subrow_column).value,)
    if ws.cell(row=row, column=location_column).value != "None":
        location_tuple += (ws.cell(row=row, column=location_column).value,)

# print(sample_number_tuple)
# print(section_tuple)
# print(subrow_tuple)
# print(location_tuple)


# 3 - defining a function to determine sections affected by a sample
# this function takes sample_number and gives the list of sections affected by the services of that sample
def find_affected_sctions(item):
    affected_sections = []
    if sample_number_tuple.index(item) == 0:  # <<< for first sample >>>
        affected_sections = [i for i in range(number_of_sections, section_tuple[0], -1)]
        affected_sections.append(section_tuple[0])

    else:  # <<< for all samples except first sample >>>
        # definitions of section and subrow cells
        # section of sample
        current_section = section_tuple[sample_number_tuple.index(item)]
        previous_section = section_tuple[sample_number_tuple.index(item) - 1]
        try:
            next_section = section_tuple[sample_number_tuple.index(item) + 1]
        except:  # for last item in section_tuple - (to handel of> IndexError: tuple index out of range)
            next_section = 27

        # Subrow of sample
        current_subrow = subrow_tuple[sample_number_tuple.index(item)]
        previous_subrow = subrow_tuple[sample_number_tuple.index(item) - 1]
        try:
            next_subrow = subrow_tuple[sample_number_tuple.index(item) + 1]
        except:
            next_subrow = current_subrow
        # print (previous_subrow,current_subrow,next_subrow)

        if current_subrow == previous_subrow:  # <<<for samples in same subrow >>>
            if current_section > previous_section:
                affected_sections = [i for i in range(previous_section, current_section + 1)]
            elif current_section < previous_section:
                affected_sections = [i for i in range(current_section, previous_section)]
            else:  # current_section = previous_section
                if current_section > number_of_sections / 2:  # right side of the pile
                    affected_sections = [i for i in range(current_section, number_of_sections + 1)]
                    print("equal_sections")
                else:  # left side of the pile
                    affected_sections = [i for i in range(1, current_section + 1)]
        else:  # <<<for samples in different subrow >>>
            if current_section < next_section:  # left side of the pile
                affected_sections = [i for i in range(1, previous_section)] \
                                    + [i for i in range(1, current_section + 1)]
            elif current_section > next_section:  # left side of the pile
                affected_sections = [i for i in range(previous_section, number_of_sections + 1)] \
                                    + [i for i in range(current_section, number_of_sections + 1)]
            else:  # current_section = next_section # could happen for last sample in pile
                affected_sections = [current_section]

        # for samples in same subrow but different direction of stacker movement:
        if len(affected_sections) < 3 and affected_sections[-1] < 5:
            affected_sections += [i for i in range(1, previous_section)]
        elif len(affected_sections) < 3 and affected_sections[-1] > 23:
            affected_sections += [i for i in range(previous_section, number_of_sections + 1)]

    return affected_sections


# 4 - define a function to extract locations for each sample
# this function takes sample_number and gives the list of locations for that sample
def find_locations_of_sample(sample_name):
    locations = location_tuple[sample_number_tuple.index(sample_name)]
    locations = locations.split(",")
    for item in locations:
        locations[locations.index(item)] = item.split(":")
    for item in locations:
        item[1] = int(item[1])  # convert type of number of services from string to int
        item[0] = item[0].replace(" ", "")  # in order to delete whitespaces in location names
    return locations


# 5 - to complete source list from data
for sample in sample_number_tuple:
    locations = find_locations_of_sample(sample)
    for item in locations:
        if item[0] not in source_list and item[0][0] != "E" and item[0][0:5] != "other":
            source_list.append(item[0])
source_list.append("Mining blocks")

# 6 - define and completion of "section-location" array >> rows are locations and columns are sections
section_location_array = np.zeros((len(source_list), number_of_sections), dtype='i')
for sample in sample_number_tuple:
    affected_sections = find_affected_sctions(sample)
    locations = find_locations_of_sample(sample)

    distributed_services = [0 for i in range(len(affected_sections))]  # just to test results

    i = 0
    for item in locations:
        source = item[0]
        service = item[1]
        while True:
            if i > (len(affected_sections)-1):  # iteration between section list items
                i = 0
            if service > 0:
                if source in source_list:
                    section_location_array[source_list.index(source)][affected_sections[i]-1] += 1
                else:
                    section_location_array[source_list.index("Mining blocks")][affected_sections[i]-1] += 1
                distributed_services[i] += 1
                service -= 1
                i += 1
            else:
                break

    print(" sample: ", sample, " section: ", section_tuple[sample_number_tuple.index(sample)],
          " affected_sections: ", affected_sections, " locations: ", locations,
          "distributed_services", distributed_services)

print("source_list", source_list)

# for item in section_location_array:
#     print (item, type(item), np.sum(item))

# 7 -   Plots ---------------------------------------------------------------:
font1 = {'family': 'serif', 'color': 'red', 'size': 15}
font2 = {'family': 'serif', 'color': 'darkred', 'size': 10}
font3 = {'family': 'cursive', 'color': 'green', 'size': 20}
font4 = {'family': 'serif', 'color': 'black', 'size': 12}

# plot #1 : image show of whole matrix of sections and services
im = plt.imshow(section_location_array, interpolation=None, cmap='RdPu')
plt.yticks([i for i in range(len(source_list))], labels=source_list, rotation=0, ha="right")
plt.xticks([i for i in range(0, 27)], labels=[i for i in range(1, 28)])
plt.title(pile_name, fontdict=font1)

for i in range(len(source_list)):
      for j in range(27):
            plt.text(j, i, str(section_location_array[i][j]), va='center', ha='center')

cbar = plt.colorbar(im)
cbar.set_label("Color bar")
plt.show()

# plot #2 : image show of  matrix of sections and services for each source of feed
for i in range(len(source_list)):
    plt.subplot(len(source_list), 1, i+1)
    sub_array=np.asarray(section_location_array[i][:])
    sub_array=sub_array.reshape(1, number_of_sections)
    plt.imshow(sub_array, cmap="RdPu")
    # print (sub_array)
    plt.xticks([i for i in range(0, 27)], labels=[i for i in range(1, 28)])
    # print ("max", max(section_location_array[i][:]))
    plt.yticks([])
    # plt.title(source_list[i],x=0.005, y=1.0, va="top", fontdict=font4, loc='left')
    cbar = plt.colorbar()
    plt.ylabel(source_list[i], rotation=0, labelpad=80, ha='left', va='top', fontdict=font2)
    for j in range(number_of_sections):
        plt.text(j, 0, str(sub_array[0][j]), va='center', ha='center')

plt.suptitle(pile_name, fontdict=font1)
plt.show()

# plot #3 : line charts of whole matrix of sections and services
for i in range (len(source_list)):
    loc_list = section_location_array[i][:]
    plt.subplot(len(source_list), 1, i+1)
    plt.plot(loc_list, label=source_list[i])
    plt.xticks([i for i in range(0, 27)], labels=[i for i in range(1, 28)])
    plt.yticks(size=8)
    # plt.title(source_list[i],x=0.005, y=0.8, va="top", fontdict=font2, loc='left')
    plt.ylabel(source_list[i], rotation=0, labelpad=70, ha='left', va='top', fontdict=font2)
    for j, v in enumerate(loc_list):
        plt.text(j, v, "%d" % v, ha="center")
    # plt.grid()

plt.suptitle(pile_name, fontdict=font1)
plt.show()
